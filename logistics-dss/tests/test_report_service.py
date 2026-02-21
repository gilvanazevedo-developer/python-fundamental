"""
Integration tests for src/services/report_service.py

Uses a clean in-memory SQLite database via the clean_database fixture.
"""

import sys
import os
from pathlib import Path
from datetime import date, timedelta
from decimal import Decimal

import pytest

PROJECT_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from src.services.report_service import ReportService
from src.database.models import Product, Warehouse, InventoryLevel, SalesRecord


# ---------------------------------------------------------------------------
# Seed helper
# ---------------------------------------------------------------------------

def _seed(db_manager):
    """Seed three products across two categories with 30 days of daily sales."""
    today = date.today()
    with db_manager.get_session() as session:
        for pid, name, cat, cost, price in [
            ("A1", "Alpha",   "CatA", "20", "40"),
            ("B1", "Beta",    "CatA",  "5", "10"),
            ("C1", "Gamma",   "CatB", "15", "25"),
        ]:
            session.add(Product(id=pid, name=name, category=cat,
                                unit_cost=Decimal(cost),
                                unit_price=Decimal(price)))

        session.add(Warehouse(id="WH1", name="Main", location="City", capacity=5000))

        for pid, qty, rev, stock in [
            ("A1", 30, 1200.0, 300),
            ("B1", 10,  100.0, 200),
            ("C1",  5,  125.0,  20),
        ]:
            session.add(InventoryLevel(product_id=pid, warehouse_id="WH1", quantity=stock))
            for offset in range(1, 31):
                session.add(SalesRecord(
                    date=today - timedelta(days=offset),
                    product_id=pid, warehouse_id="WH1",
                    quantity_sold=qty,
                    revenue=Decimal(str(rev)),
                ))
        session.commit()


# ---------------------------------------------------------------------------
# get_executive_report
# ---------------------------------------------------------------------------

class TestGetExecutiveReport:

    def test_returns_dict(self, clean_database):
        _seed(clean_database)
        svc = ReportService(db_manager=clean_database)
        result = svc.get_executive_report(days=30)
        assert isinstance(result, dict)

    def test_required_top_level_keys(self, clean_database):
        _seed(clean_database)
        svc = ReportService(db_manager=clean_database)
        result = svc.get_executive_report(days=30)
        required = {
            "period_days", "generated_at",
            "financial_summary", "stock_health", "service_level",
            "abc_summary", "top_products", "sales_trend",
            "reorder_alerts", "optimization_summary",
            "top_savings_skus", "stock_by_category",
        }
        assert required.issubset(result.keys())

    def test_period_days_set_correctly(self, clean_database):
        svc = ReportService(db_manager=clean_database)
        result = svc.get_executive_report(days=60)
        assert result["period_days"] == 60

    def test_abc_summary_has_three_classes(self, clean_database):
        _seed(clean_database)
        svc = ReportService(db_manager=clean_database)
        result = svc.get_executive_report(days=30)
        classes = [s["abc_class"] for s in result["abc_summary"]]
        assert set(classes) == {"A", "B", "C"}

    def test_top_products_is_list(self, clean_database):
        _seed(clean_database)
        svc = ReportService(db_manager=clean_database)
        result = svc.get_executive_report(days=30)
        assert isinstance(result["top_products"], list)

    def test_sales_trend_has_entries(self, clean_database):
        _seed(clean_database)
        svc = ReportService(db_manager=clean_database)
        result = svc.get_executive_report(days=30)
        assert len(result["sales_trend"]) > 0

    def test_reorder_alerts_filtered_to_critical_warning(self, clean_database):
        _seed(clean_database)
        svc = ReportService(db_manager=clean_database)
        result = svc.get_executive_report(days=30)
        for alert in result["reorder_alerts"]:
            assert alert["urgency"] in ("CRITICAL", "WARNING")

    def test_optimization_summary_keys(self, clean_database):
        _seed(clean_database)
        svc = ReportService(db_manager=clean_database)
        result = svc.get_executive_report(days=30)
        opt = result["optimization_summary"]
        for key in ("total_savings", "savings_pct", "total_products"):
            assert key in opt

    def test_top_savings_at_most_10_rows(self, clean_database):
        _seed(clean_database)
        svc = ReportService(db_manager=clean_database)
        result = svc.get_executive_report(days=30)
        assert len(result["top_savings_skus"]) <= 10

    def test_empty_database_no_crash(self, clean_database):
        svc = ReportService(db_manager=clean_database)
        result = svc.get_executive_report(days=30)
        assert "period_days" in result


# ---------------------------------------------------------------------------
# export_to_excel
# ---------------------------------------------------------------------------

class TestExportToExcel:

    def test_creates_file(self, clean_database, tmp_path):
        _seed(clean_database)
        svc = ReportService(db_manager=clean_database)
        report = svc.get_executive_report(days=30)
        filepath = str(tmp_path / "report.xlsx")
        result = svc.export_to_excel(report, filepath)
        assert result is True
        assert os.path.exists(filepath)

    def test_file_has_expected_sheets(self, clean_database, tmp_path):
        _seed(clean_database)
        svc = ReportService(db_manager=clean_database)
        report = svc.get_executive_report(days=30)
        filepath = str(tmp_path / "report.xlsx")
        svc.export_to_excel(report, filepath)

        import openpyxl
        wb = openpyxl.load_workbook(filepath)
        expected_sheets = {
            "Summary", "ABC Analysis", "Top Products",
            "Sales Trend", "Reorder Alerts", "Optimization"
        }
        assert expected_sheets.issubset(set(wb.sheetnames))

    def test_summary_sheet_has_data(self, clean_database, tmp_path):
        _seed(clean_database)
        svc = ReportService(db_manager=clean_database)
        report = svc.get_executive_report(days=30)
        filepath = str(tmp_path / "report.xlsx")
        svc.export_to_excel(report, filepath)

        import openpyxl
        wb = openpyxl.load_workbook(filepath)
        ws = wb["Summary"]
        assert ws.max_row > 1  # has header + data rows

    def test_returns_false_on_invalid_path(self, clean_database):
        svc = ReportService(db_manager=clean_database)
        report = {"period_days": 30}
        result = svc.export_to_excel(report, "/nonexistent/path/report.xlsx")
        assert result is False

    def test_empty_report_still_creates_file(self, clean_database, tmp_path):
        svc = ReportService(db_manager=clean_database)
        report = svc.get_executive_report(days=30)  # empty DB
        filepath = str(tmp_path / "empty.xlsx")
        result = svc.export_to_excel(report, filepath)
        assert result is True


# ---------------------------------------------------------------------------
# export_to_csv
# ---------------------------------------------------------------------------

class TestExportToCsv:

    def test_creates_folder_and_files(self, clean_database, tmp_path):
        _seed(clean_database)
        svc = ReportService(db_manager=clean_database)
        report = svc.get_executive_report(days=30)
        folder = str(tmp_path / "csv_export")
        result = svc.export_to_csv(report, folder)
        assert result is True
        assert os.path.isdir(folder)

    def test_expected_csv_files_created(self, clean_database, tmp_path):
        _seed(clean_database)
        svc = ReportService(db_manager=clean_database)
        report = svc.get_executive_report(days=30)
        folder = tmp_path / "csv_export"
        svc.export_to_csv(report, str(folder))
        expected = [
            "executive_summary.csv", "abc_analysis.csv",
            "top_products.csv", "sales_trend.csv",
        ]
        for filename in expected:
            assert (folder / filename).exists(), f"Missing: {filename}"

    def test_csv_files_have_content(self, clean_database, tmp_path):
        _seed(clean_database)
        svc = ReportService(db_manager=clean_database)
        report = svc.get_executive_report(days=30)
        folder = tmp_path / "csv_export"
        svc.export_to_csv(report, str(folder))

        import csv as csvlib
        with open(folder / "sales_trend.csv", newline="") as f:
            rows = list(csvlib.DictReader(f))
        assert len(rows) > 0

    def test_returns_false_on_invalid_folder(self, clean_database):
        svc = ReportService(db_manager=clean_database)
        # Write to a path where a file blocks directory creation
        # Use a non-writable path simulation by passing a file as folder
        import tempfile
        with tempfile.NamedTemporaryFile() as tf:
            result = svc.export_to_csv({"abc_summary": [{"x": 1}]}, tf.name)
            # tf.name is a file, not a directory → mkdir raises error → False
            assert result is False

    def test_returns_true_on_success(self, clean_database, tmp_path):
        svc = ReportService(db_manager=clean_database)
        report = svc.get_executive_report(days=30)
        result = svc.export_to_csv(report, str(tmp_path / "out"))
        assert result is True
